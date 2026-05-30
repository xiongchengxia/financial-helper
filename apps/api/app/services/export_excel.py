from io import BytesIO

from openpyxl import Workbook

from app.models import DocumentRecord, VoucherRecord


def build_vouchers_workbook(
    vouchers: list[VoucherRecord],
    documents_by_id: dict[str, DocumentRecord],
) -> bytes:
    wb = Workbook()
    ws_header = wb.active
    ws_header.title = "凭证主表"
    ws_header.append(
        [
            "凭证ID",
            "凭证日期",
            "字",
            "号",
            "附件张数",
            "借方合计",
            "贷方合计",
            "来源文档ID",
            "发票号码",
            "合规是否通过",
        ]
    )

    ws_lines = wb.create_sheet("分录明细")
    ws_lines.append(["凭证ID", "行号", "科目编码", "科目名称", "借方金额", "贷方金额", "摘要"])

    ws_risks = wb.create_sheet("风险明细")
    ws_risks.append(["文档ID", "风险码", "级别", "说明", "证据"])

    for v in vouchers:
        doc_id = v.documentIds[0] if v.documentIds else ""
        doc = documents_by_id.get(doc_id)
        invoice_no = ""
        compliance_passed = ""
        if doc:
            invoice_no = doc.payload.get("invoice_number", "")
            if doc.compliance:
                compliance_passed = "是" if doc.compliance.passed else "否"
            for risk in doc.risks:
                ws_risks.append(
                    [
                        doc.id,
                        risk.code,
                        risk.level.value,
                        risk.message,
                        str(risk.evidence),
                    ]
                )

        ws_header.append(
            [
                v.id,
                v.voucherDate,
                v.voucherWord,
                v.voucherNo,
                v.attachmentCount,
                v.debitTotal,
                v.creditTotal,
                doc_id,
                invoice_no,
                compliance_passed,
            ]
        )
        for line in v.lines:
            ws_lines.append(
                [
                    v.id,
                    line.lineNo,
                    line.accountCode,
                    line.accountName,
                    line.debit,
                    line.credit,
                    line.summary,
                ]
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
